import sys
from tqdm import tqdm
import time
import os
import gspread
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

def show_loading(duration=5):
        """
        Displays a console-based progress bar before input.
        """
        print("\n\nInitializing program, please wait...")
        for _ in tqdm(range(duration), desc="Loading", ncols=75):
            time.sleep(1)  # Simulates loading process
            
def list_available_sheets(credentials_file):
    """
    Fetches all available Google Sheet names under the authenticated account.
    """
    try:
        gc = gspread.service_account(filename=credentials_file)
        return [sheet.title for sheet in gc.openall()]  # Get list of sheet titles
    except Exception as e:
        print(f"Error accessing Google Sheets: {e}")
        return None

def import_google_sheet(credentials_file, sheet_name, day_name):
    """
    Imports data from a Google Sheet into a Pandas DataFrame.
    """
    try:
        gc = gspread.service_account(filename=credentials_file)
        sheet = gc.open(sheet_name).worksheet(day_name)
        data = sheet.get_all_records()
        return pd.DataFrame(data)
    except Exception as e:
        print(f"Error importing Google Sheet: {e}")
        return None

def process_data(df):
    """
    Processes student data to determine their next level.
    """
    swimmer_progression = {
        "Adult 1": "Adult 2",
        "Adult 2": "Adult 3",
        "Adult 3": "Adult Fitness",
        "Teen 1": "Teen 2",
        "Teen 2": "Teen 3",
        "Teen 3": "Teen Fitness",
        "Preschool 1": "Swimmer 1",
        "Preschool 2": "Swimmer 1",
        "Preschool 3": "Swimmer 1",
        "Preschool 4": "Swimmer 2",
        "Preschool 5": {1: "Swimmer 3", 0: "Swimmer 2"},
        "Swimmer 1": "Swimmer 2",
        "Swimmer 2": "Swimmer 3",
        "Swimmer 3": "Swimmer 4",
        "Swimmer 4": "Swimmer 5",
        "Swimmer 5": "Swimmer 6",
        "Swimmer 6": "Swim Patrol Rookie",
        "Swim Patrol Rookie": "Swim Patrol Ranger",
        "Swim Patrol Ranger": "Swim Patrol Star",
        "Swim Patrol Star": "Bronze Medallion",
        "Private": "Private",
        "Adapted Private": "Adapted Private",
    }
    def determine_next_level(row):
            current_level = row["Current Level"]
            pass_fail = int(row["Pass/Fail"])
            age = int(row["Age"])  # Ensure Age is an integer

            #If swimmer is 5 years old, convert Preschool level to Swimmer equivalent
            preschool_to_swimmer = {
                "Preschool 1": "Swimmer 1",
                "Preschool 2": "Swimmer 1",
                "Preschool 3": "Swimmer 2" if pass_fail == 1 else "Swimmer 1",
                "Preschool 4": "Swimmer 2",
                "Preschool 5": "Swimmer 3" if pass_fail == 1 else "Swimmer 2"
            }

            if age == 5 and current_level in preschool_to_swimmer:
                return preschool_to_swimmer[current_level]
            if current_level=="Parent & Tot 1" and age==1:
                return "Parent & Tot 2"
            if current_level=="Parent & Tot 2" and age==2:
                return "Parent & Tot 3"
            #If Parent & Tot swimmer is older than 3, move them to Preschool 1
            if current_level.startswith("Parent & Tot") and age > 3:
                return "Preschool 1"

            # Handle Preschool 5 separately (because of nested dictionary)
            if current_level == "Preschool 5":
                return swimmer_progression[current_level].get(pass_fail, "Swimmer 2")

            if pass_fail == 1 and current_level in swimmer_progression:
                return swimmer_progression[current_level]

            return current_level  # Stay in the same level if failed

         #Fix indentation here!
    df["Current Level"] = df["Current Level"].str.replace("Ed: ", "", regex=False)

    df["Pass/Fail"] = df["Pass/Fail"].astype(int)
    df["Age"] = pd.to_numeric(df["Age"], errors="coerce").fillna(0).astype(int)

    def extract_first_name(monitor):
        parts = monitor.split(", ")

        if len(parts) > 3:
            return f"{parts[1]} & {parts[3]}"  

        if len(parts) > 1:
            return parts[1]  

        return monitor  

    df["Instructor Name"] = df["Monitor"].apply(extract_first_name)
    df["Next Level"] = df.apply(determine_next_level, axis=1)

    return df
        
def export_to_template_excel(df, template_path, output_path):
    """
    Exports processed data to a formatted Excel template.
    """
    workbook = load_workbook(template_path)
    template_sheet_name = workbook.sheetnames[0]
    cell_positions = [
        {"name": "C8", "level": "D10", "instructor": "D12"},
        {"name": "C26", "level": "D28", "instructor": "D30"},
        {"name": "C44", "level": "D46", "instructor": "D48"},
        {"name": "C62", "level": "D64", "instructor": "D66"},
    ]
    sheet_counter, row_counter = 1, 0
    grouped = df.groupby("Instructor Name")
    for instructor, group in grouped:
        for _, row in group.iterrows():
            if row_counter == 0:
                sheet_name = f"Sheet {sheet_counter}"
                workbook.copy_worksheet(workbook[template_sheet_name]).title = sheet_name
                current_sheet = workbook[sheet_name]
                sheet_counter += 1
            position = cell_positions[row_counter]
            current_sheet[position["name"]] = row["Name"]
            current_sheet[position["level"]] = row["Next Level"]
            current_sheet[position["instructor"]] = instructor
            row_counter = (row_counter + 1) % 4
    del workbook[template_sheet_name]
    workbook.save(output_path)
    print(f"Data successfully exported to {output_path}")


def get_user_inputs():
    """
    Opens file dialogs for user to select Google Sheet and Excel paths.
    """
    root = tk.Tk()
    root.withdraw()
    

    while True:
        sheet_name = input("\nEnter the Google Sheet name: ")
        df = list_available_sheets(credentials_file)
        if sheet_name == '0':
            sys.exit()
        if sheet_name in df:
            break  #Exit loop when a valid sheet is found
        else:
            print("Invalid Google Sheet name. Please try again.")

    print("\nGoogle Sheet loaded successfully!")
    # Call the loading bar before first user input
    while True:  # ✅ Allow rerunning the process without restarting the program
        print("\nSelect the day for processing:")
        day = {'1': "Wednesday", '2': "Friday", '3': "Saturday", '4': "Sunday"}
        day_name = input("\n\nSelect the Day - "
                         "\n1. Wednesday"
                         "\n2. Friday"
                         "\n3. Saturday"
                         "\n4. Sunday"
                         "\n0. Exit Program"
                         "\nInput: ")

        if day_name == '0':
            print("Exiting program...")
            sys.exit()  # ✅ Exit program completely

        if day_name in day:
            day_name = day[day_name]
            print(f"Selecting the shift for {day_name}")

        elif day_name =="5":
            day_name = "Other"
        else:
            print("Invalid selection. Please try again.")
            continue  # Restart loop if input is invalid

        while True:  #Loop back if user cancels file selection
            template_file = filedialog.askopenfilename(title="Select Excel Template")
            if not template_file:
                print("File selection canceled. Returning to day selection.")
                break  # Goes back to selecting day_name

            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save Processed Excel File")
            if not output_file:
                print("Save file selection canceled. Returning to day selection.")
                break  # Goes back to selecting day_name

            return sheet_name, day_name, template_file, output_file  #Successfully selected all inputs


if __name__ == "__main__":
    show_loading()
    #Make sure this is changed to the correct format and is hardcoded. If not, get new and save as .json
    credentials_file=r"C:\Users\maste\Desktop\PreregProject\civic-reserve-447800-h5-0ea4d7319ace.json"
    sheet_name, day_name, template_file, output_file = get_user_inputs()

    print("\nImporting data from Google Sheet...")
    df = import_google_sheet(credentials_file, sheet_name, day_name)
    
    if df is not None:
        print("Processing data...")
        processed_data = process_data(df)
        print("Exporting to Excel template...")
        export_to_template_excel(processed_data, template_file, output_file)
        print("Process completed successfully!")
    while True:
        sheet_name, template_file, output_file = get_user_inputs()

        print(f"\nProcessing data for {day_name}...")
        # Add function calls to process the data and export it (not shown here)


