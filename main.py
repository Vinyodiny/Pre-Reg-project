import gspread
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

def import_google_sheet(credentials_file, sheet_name):
    """
    Imports data from a Google Sheet into a Pandas DataFrame.
    """
    try:
        gc = gspread.service_account(filename=credentials_file)
        sheet = gc.open(sheet_name).sheet1
        data = sheet.get_all_records()
        return pd.DataFrame(data)
    except Exception as e:
        print(f"Error importing Google Sheet: {e}")
        return None

def process_data(df):
    """
    Processes student data to determine their next level.
    """
    swimmer_progression = {
        "Preschool 1": "Swimmer 1",
        "Preschool 2": "Swimmer 1",
        "Preschool 3": "Swimmer 1",
        "Preschool 4": "Swimmer 2",
        "Preschool 5": {1: "Swimmer 3", 0: "Swimmer 2"},
        "Swimmer 1": "Swimmer 2",
        "Swimmer 2": "Swimmer 3",
        "Swimmer 3": "Swimmer 4",
        "Swimmer 4": "Swimmer 5",
        "Swimmer 5": "Swimmer 6",
    }
    def determine_next_level(row):
        current_level = row["Current Level"]
        pass_fail = row["Pass/Fail"]
        if current_level == "Preschool 5":
            return swimmer_progression[current_level][pass_fail]
        return swimmer_progression.get(current_level, current_level)
    df["Next Level"] = df.apply(determine_next_level, axis=1)
    df["Instructor Name"] = df["Monitor"].apply(lambda x: x.split(", ")[1] if ", " in x else x)
    return df

def export_to_template_excel(df, template_path, output_path):
    """
    Exports processed data to a formatted Excel template.
    """
    workbook = load_workbook(template_path)
    template_sheet_name = workbook.sheetnames[0]
    cell_positions = [
        {"name": "C8", "level": "D10", "instructor": "D12"},
        {"name": "C26", "level": "D28", "instructor": "D30"},
        {"name": "C44", "level": "D46", "instructor": "D48"},
        {"name": "C62", "level": "D64", "instructor": "D66"},
    ]
    sheet_counter, row_counter = 1, 0
    grouped = df.groupby("Instructor Name")
    for instructor, group in grouped:
        for _, row in group.iterrows():
            if row_counter == 0:
                sheet_name = f"Sheet {sheet_counter}"
                workbook.copy_worksheet(workbook[template_sheet_name]).title = sheet_name
                current_sheet = workbook[sheet_name]
                sheet_counter += 1
            position = cell_positions[row_counter]
            current_sheet[position["name"]] = row["Name"]
            current_sheet[position["level"]] = row["Next Level"]
            current_sheet[position["instructor"]] = instructor
            row_counter = (row_counter + 1) % 4
    del workbook[template_sheet_name]
    workbook.save(output_path)
    print(f"Data successfully exported to {output_path}")

def get_user_inputs():
    """
    Opens file dialogs for user to select Google Sheet and Excel paths.
    """
    root = tk.Tk()
    root.withdraw()
    credentials_file = filedialog.askopenfilename(title="Select Google API Credentials JSON")
    sheet_name = input("Enter Google Sheet Name: ")
    template_file = filedialog.askopenfilename(title="Select Excel Template")
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Save Processed Excel File")
    return credentials_file, sheet_name, template_file, output_file

if __name__ == "__main__":
    credentials_file, sheet_name, template_file, output_file = get_user_inputs()
    print("Importing data from Google Sheet...")
    df = import_google_sheet(credentials_file, sheet_name)
    if df is not None:
        print("Processing data...")
        processed_data = process_data(df)
        print("Exporting to Excel template...")
        export_to_template_excel(processed_data, template_file, output_file)
        print("Process completed successfully!")
